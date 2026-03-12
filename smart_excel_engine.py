"""
ExcelMind Smart Excel Engine
=============================
Backend motor: Ham Excel yüklenir → otomatik analiz → Akıllı Excel üretilir.

Fonksiyonlar:
  - detect_data_type(headers, data_rows)  → Veri türü algılama
  - analyze_for_dashboard(input_path)      → Dashboard için JSON analiz
  - build_smart_excel(input_path, output)  → Akıllı Excel üretimi
"""

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule, DataBarRule
from openpyxl.chart import BarChart, PieChart, LineChart, Reference
from openpyxl.chart.label import DataLabelList
from datetime import datetime
import json
import os

# ============================================================
# STYLES
# ============================================================
BLUE = "2F80ED"; BLUE_DARK = "1A5276"; GREEN = "27AE60"
ORANGE = "E67E22"; PURPLE = "9B59B6"; RED = "E74C3C"

hdr_fill = PatternFill("solid", fgColor=BLUE_DARK)
blue_fill = PatternFill("solid", fgColor=BLUE)
green_fill = PatternFill("solid", fgColor="EAFAF1")
light_fill = PatternFill("solid", fgColor="F8FAFC")
filter_fill = PatternFill("solid", fgColor="EEF5FF")
alt_fill = PatternFill("solid", fgColor="FAFBFC")

hdr_font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
title_font = Font(name="Arial", bold=True, color=BLUE_DARK, size=16)
sub_title = Font(name="Arial", italic=True, color="6B7C93", size=10)
label_font = Font(name="Arial", bold=True, color=BLUE_DARK, size=10)
value_font = Font(name="Arial", bold=True, color="1A2332", size=14)
normal_font = Font(name="Arial", size=10)
bold_sm = Font(name="Arial", bold=True, size=10)
thin_border = Border(
    left=Side('thin', color='D0D5DD'), right=Side('thin', color='D0D5DD'),
    top=Side('thin', color='D0D5DD'), bottom=Side('thin', color='D0D5DD')
)

MONTH_NAMES = ['Ocak','Şubat','Mart','Nisan','Mayıs','Haziran',
               'Temmuz','Ağustos','Eylül','Ekim','Kasım','Aralık']

# ============================================================
# DATA TYPE DETECTION
# ============================================================
def detect_data_type(headers, data_rows):
    """
    Detect data type from headers AND first 30 rows of data values.
    Returns: 'sales', 'ecommerce', 'inventory', 'bank', 'finance', 'cashflow', 'hr', 'customers', 'general'
    """
    h = [str(x).lower().replace('_','').replace('-','').replace(' ','') for x in headers if x]
    header_text = ' '.join(h)

    # Also scan data values
    sample_values = []
    for row in data_rows[:30]:
        for val in row:
            if val is not None:
                sample_values.append(str(val).lower())
    data_text = ' '.join(sample_values)
    all_text = header_text + ' ' + data_text

    # Sales (value/amount + product/customer columns in headers)
    import re
    if re.search(r'sati[sş]|sale|revenue|ciro|fatura|invoice|tutar|fiyat|amount|adet', header_text) and \
       re.search(r'ürün|urun|product|müşteri|musteri|customer|quantity', header_text):
        return 'sales'

    # E-commerce
    if re.search(r'sipari[sş]|order|sku|marketplace|trendyol|amazon|hepsiburada|shopify', header_text):
        return 'ecommerce'

    # Inventory
    if re.search(r'stok|stock|envanter|inventory|depo|warehouse|miktar', header_text):
        return 'inventory'

    # Bank
    if re.search(r'banka|bank|iban|havale|eft|virman', header_text):
        return 'bank'
    # Bank also detectable by "hesap" + transaction patterns
    if re.search(r'hesap', header_text) and re.search(r'bakiye|tutar|işlem', header_text):
        return 'bank'

    # HR — check BEFORE finance (maaş/salary can overlap, but departman+pozisyon is unique to HR)
    if re.search(r'departman|department|pozisyon|position|çalışan|calisan|employee', header_text):
        return 'hr'
    if re.search(r'maaş|maas|salary', header_text) and re.search(r'personel|sicil|işegiriş|isegiris|unvan', header_text):
        return 'hr'

    # Finance (Gelir-Gider) — headers first, then data values with specific patterns (avoid "kar" matching names)
    if re.search(r'gelir|gider|expense|income|maliyet|cost|kâr|profit|zarar|loss|bütçe|butce|budget', header_text):
        return 'finance'
    if re.search(r'\bgelir\b|\bgider\b', data_text):
        return 'finance'

    # Cash flow
    if re.search(r'nakit|cash|flow|tahsilat|borç|borc|alacak', header_text):
        return 'cashflow'

    # HR fallback (maaş alone)
    if re.search(r'maaş|maas|salary', header_text):
        return 'hr'

    # Customer
    if re.search(r'müşteri|musteri|customer|client|firma|company|telefon|phone|email|adres|address', header_text):
        return 'customers'

    return 'general'

# ============================================================
# READ EXCEL DATA
# ============================================================
def read_excel(input_path):
    """Read Excel file, return (headers, data_rows, sheet_name)"""
    src = openpyxl.load_workbook(input_path, data_only=True)
    ws = src.active
    sheet_name = ws.title or "Veri"
    headers = [cell.value for cell in ws[1]]
    data_rows = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        if any(v is not None for v in row):
            row_list = list(row)
            # Parse date strings (DD.MM.YYYY format)
            for ci in range(len(row_list)):
                if row_list[ci] and isinstance(row_list[ci], str):
                    try:
                        parts = row_list[ci].split('.')
                        if len(parts) == 3 and len(parts[2]) == 4:
                            row_list[ci] = datetime(int(parts[2]), int(parts[1]), int(parts[0]))
                    except:
                        pass
            data_rows.append(row_list)

    src.close()
    return headers, data_rows, sheet_name

# ============================================================
# COLUMN ANALYSIS
# ============================================================
def analyze_columns(headers, data_rows):
    """Detect column roles: date, numeric, categorical, text"""
    roles = {}
    n = len(data_rows)

    for ci, header in enumerate(headers):
        if header is None:
            continue
        vals = [r[ci] for r in data_rows if ci < len(r) and r[ci] is not None]
        if not vals:
            continue

        # Date
        if any(isinstance(v, datetime) for v in vals[:20]):
            roles[ci] = {'type': 'date', 'header': header}
            continue

        # Numeric
        numeric_count = sum(1 for v in vals[:50] if isinstance(v, (int, float)))
        if numeric_count > len(vals[:50]) * 0.7:
            roles[ci] = {'type': 'numeric', 'header': header}
            continue

        # Categorical
        str_vals = [str(v).strip() for v in vals if v]
        unique = set(str_vals)
        if 2 <= len(unique) <= 30 and len(unique) < n * 0.3:
            roles[ci] = {'type': 'categorical', 'header': header, 'unique': sorted(unique)}
            continue

        # Text (high cardinality)
        roles[ci] = {'type': 'text', 'header': header}

    return roles


def pick_main_value(roles, headers, data_rows):
    """Pick the best numeric column as main value"""
    value_cols = [ci for ci, info in roles.items() if info['type'] == 'numeric']

    if not value_cols:
        return None

    # Keyword match (prefer total/aggregate columns over unit price)
    # Priority 1: "toplam" keywords (aggregate columns)
    priority_keywords = ['toplam', 'total', 'net', 'brüt', 'brut']
    # Priority 2: general value keywords
    value_keywords = ['tutar', 'fiyat', 'maaş', 'maas', 'amount',
                      'gelir', 'gider', 'satış', 'satis', 'revenue', 'price',
                      'maliyet', 'cost', 'prim', 'ücret', 'ucret']
    avoid_keywords = ['bakiye', 'balance', 'sıra', 'no', 'id', 'numara', 'birim']

    # First try priority keywords
    for vc in value_cols:
        h = str(headers[vc]).lower()
        if any(kw in h for kw in priority_keywords) and not any(kw in h for kw in avoid_keywords):
            return vc

    # Then try general value keywords
    for vc in value_cols:
        h = str(headers[vc]).lower()
        if any(kw in h for kw in value_keywords) and not any(kw in h for kw in avoid_keywords):
            return vc

    # Coefficient of variation fallback (higher = more likely transaction amount, not running total)
    filtered = [vc for vc in value_cols if not any(kw in str(headers[vc]).lower() for kw in avoid_keywords)]
    if not filtered:
        filtered = value_cols

    best_cv = -1
    best_col = filtered[0]
    for vc in filtered:
        vals = [r[vc] for r in data_rows[:100] if vc < len(r) and isinstance(r[vc], (int, float))]
        if len(vals) < 5:
            continue
        avg = sum(vals) / len(vals)
        if avg == 0:
            continue
        variance = sum((v - avg) ** 2 for v in vals) / len(vals)
        cv = (variance ** 0.5) / abs(avg)
        if cv > best_cv:
            best_cv = cv
            best_col = vc

    return best_col


def pick_roles(roles, headers, data_rows):
    """Pick: date_col, main_value, value_cols, filter_cols (up to 4)"""
    date_col = None
    filter_cols = []

    for ci, info in roles.items():
        if info['type'] == 'date' and date_col is None:
            date_col = ci
        elif info['type'] == 'categorical':
            filter_cols.append(ci)

    value_cols = [ci for ci, info in roles.items() if info['type'] == 'numeric']
    main_value = pick_main_value(roles, headers, data_rows)

    return date_col, main_value, value_cols, filter_cols[:4]

# ============================================================
# ANALYZE FOR DASHBOARD (JSON response)
# ============================================================
def analyze_for_dashboard(input_path, data_type='auto'):
    """Analyze Excel file and return JSON-serializable dict for frontend dashboard"""
    headers, data_rows, sheet_name = read_excel(input_path)

    if data_type == 'auto':
        data_type = detect_data_type(headers, data_rows)

    roles = analyze_columns(headers, data_rows)
    date_col, main_value, value_cols, filter_cols = pick_roles(roles, headers, data_rows)

    if main_value is None:
        return {'error': 'No numeric value column found', 'headers': [str(h) for h in headers]}

    # Basic stats
    num_values = [r[main_value] for r in data_rows if main_value < len(r) and isinstance(r[main_value], (int, float))]
    total = sum(num_values)
    avg = total / len(num_values) if num_values else 0
    max_val = max(num_values) if num_values else 0

    # Category breakdown
    cat_col = filter_cols[0] if filter_cols else None
    cat_breakdown = {}
    if cat_col is not None:
        for r in data_rows:
            key = str(r[cat_col] if cat_col < len(r) else 'Diğer').strip()
            if not key:
                key = 'Diğer'
            if key not in cat_breakdown:
                cat_breakdown[key] = {'count': 0, 'total': 0}
            cat_breakdown[key]['count'] += 1
            val = r[main_value] if main_value < len(r) and isinstance(r[main_value], (int, float)) else 0
            cat_breakdown[key]['total'] += val

    # Monthly trend
    monthly = {}
    if date_col is not None:
        for r in data_rows:
            d = r[date_col] if date_col < len(r) else None
            if isinstance(d, datetime):
                key = f"{d.year}-{d.month:02d}"
                label = f"{d.year} {MONTH_NAMES[d.month-1]}"
                if key not in monthly:
                    monthly[key] = {'label': label, 'total': 0, 'count': 0}
                val = r[main_value] if main_value < len(r) and isinstance(r[main_value], (int, float)) else 0
                monthly[key]['total'] += val
                monthly[key]['count'] += 1

    return {
        'type': data_type,
        'sheet_name': sheet_name,
        'rows': len(data_rows),
        'columns': len(headers),
        'headers': [str(h) for h in headers if h],
        'main_value_col': str(headers[main_value]),
        'category_col': str(headers[cat_col]) if cat_col is not None else None,
        'date_col': str(headers[date_col]) if date_col is not None else None,
        'filter_cols': [str(headers[fc]) for fc in filter_cols],
        'stats': {
            'total': total,
            'count': len(num_values),
            'average': avg,
            'max': max_val,
        },
        'category_breakdown': cat_breakdown,
        'monthly_trend': [monthly[k] for k in sorted(monthly.keys())],
    }

# ============================================================
# BUILD SMART EXCEL
# ============================================================
def build_smart_excel(input_path, output_path):
    """
    Main function: Read raw Excel → Generate Smart Excel with:
    - Dashboard with dropdown filters
    - KPI section with SUMIFS formulas
    - Pivot tables with IF(OR()) wrappers
    - Charts (bar, pie, line)
    - Volatile helper cells for auto-recalculation
    - Monthly trend sheet
    - Hidden _Listeler sheet for DataValidation

    Returns stats dict.
    """
    headers, data_rows, sheet_name = read_excel(input_path)
    roles = analyze_columns(headers, data_rows)
    date_col, main_value, value_cols, filter_cols = pick_roles(roles, headers, data_rows)

    n_rows = len(data_rows)
    de = n_rows + 1  # data end row in Ham Veri

    print(f"[ENGINE] Data: {n_rows} rows, {len(headers)} columns")
    print(f"[ENGINE] Date: {headers[date_col] if date_col is not None else 'None'}")
    print(f"[ENGINE] Main Value: {headers[main_value] if main_value is not None else 'None'}")
    print(f"[ENGINE] Filters: {[headers[f] for f in filter_cols]}")

    if main_value is None:
        raise ValueError("No numeric value column found in the data!")

    # Extract unique values for filters
    filter_data = {}
    for fi in filter_cols:
        vals = sorted(set(str(r[fi]).strip() for r in data_rows if fi < len(r) and r[fi] is not None))
        filter_data[fi] = vals

    # Extract months
    months_set = set()
    if date_col is not None:
        for r in data_rows:
            d = r[date_col] if date_col < len(r) else None
            if isinstance(d, datetime):
                months_set.add(f"{d.year} {MONTH_NAMES[d.month-1]}")
    months_list = sorted(months_set)

    # ==========================================
    # CREATE WORKBOOK
    # ==========================================
    wb = Workbook()

    # --- Hidden _Listeler sheet ---
    ws_lists = wb.create_sheet("_Listeler")
    ws_lists.sheet_state = 'hidden'

    list_refs = {}
    list_col = 1
    for fi in filter_cols:
        ws_lists.cell(row=1, column=list_col, value="Tümü")
        for i, val in enumerate(filter_data[fi]):
            ws_lists.cell(row=2 + i, column=list_col, value=val)
        list_refs[fi] = (list_col, 1 + len(filter_data[fi]))
        list_col += 1

    # Months list
    month_list_ref = None
    if months_list:
        ws_lists.cell(row=1, column=list_col, value="Tümü")
        for i, m in enumerate(months_list):
            ws_lists.cell(row=2 + i, column=list_col, value=m)
        month_list_ref = (list_col, 1 + len(months_list))
        list_col += 1

    # --- Ham Veri sheet ---
    ws_raw = wb.create_sheet("Ham Veri")
    ws_raw.sheet_properties.tabColor = "8899AA"

    for ci, h in enumerate(headers):
        c = ws_raw.cell(row=1, column=ci + 1, value=h)
        c.font = hdr_font
        c.fill = hdr_fill
        c.border = thin_border
        c.alignment = Alignment(horizontal='center')

    for ri, row_data in enumerate(data_rows):
        for ci, val in enumerate(row_data):
            c = ws_raw.cell(row=ri + 2, column=ci + 1, value=val)
            c.font = normal_font
            c.border = thin_border
            if isinstance(val, datetime):
                c.number_format = 'DD.MM.YYYY'
            elif isinstance(val, (int, float)):
                c.number_format = '#,##0'

    ws_raw.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{de}"
    for ci in range(len(headers)):
        ws_raw.column_dimensions[get_column_letter(ci + 1)].width = max(len(str(headers[ci] or '')) + 4, 14)

    # Month helper column in Ham Veri
    month_helper_col = len(headers) + 1
    if date_col is not None:
        date_xl_col = get_column_letter(date_col + 1)
        ws_raw.cell(row=1, column=month_helper_col, value="Ay").font = hdr_font
        ws_raw.cell(row=1, column=month_helper_col).fill = hdr_fill
        ws_raw.cell(row=1, column=month_helper_col).border = thin_border
        for ri in range(n_rows):
            r = ri + 2
            formula = (f'=IF({date_xl_col}{r}="","",YEAR({date_xl_col}{r})&" "'
                       f'&CHOOSE(MONTH({date_xl_col}{r}),'
                       f'"Ocak","Şubat","Mart","Nisan","Mayıs","Haziran",'
                       f'"Temmuz","Ağustos","Eylül","Ekim","Kasım","Aralık"))')
            ws_raw.cell(row=r, column=month_helper_col, value=formula).font = normal_font
        ws_raw.column_dimensions[get_column_letter(month_helper_col)].width = 18

    # --- DASHBOARD sheet ---
    ws = wb.active
    ws.title = "Dashboard"
    ws.sheet_properties.tabColor = BLUE

    ws.merge_cells('A1:H1')
    ws['A1'] = f"ExcelMind — {sheet_name} Dashboard"
    ws['A1'].font = Font(name="Arial", bold=True, color=BLUE, size=20)
    ws['A1'].alignment = Alignment(vertical='center')
    ws.row_dimensions[1].height = 45

    ws.merge_cells('A2:H2')
    ws['A2'] = "Dropdown filtrelerden seçim yapın → KPI'lar, tablolar ve grafikler otomatik güncellenir"
    ws['A2'].font = sub_title

    # FILTERS (Row 4-6)
    r = 4
    for c in range(1, 9):
        ws.cell(row=r, column=c).fill = hdr_fill
        ws.cell(row=r, column=c).font = hdr_font
    ws.cell(row=r, column=1, value="FİLTRE")
    ws.cell(row=r, column=2, value="SEÇİM")
    ws.cell(row=r, column=4, value="FİLTRE")
    ws.cell(row=r, column=5, value="SEÇİM")
    ws.row_dimensions[r].height = 28

    # Place filters
    filter_cells = []
    positions = [(5, 1, 2), (6, 1, 2), (5, 4, 5), (6, 4, 5)]

    active_filters = filter_cols[:4]
    has_month_filter = date_col is not None and len(active_filters) < 4

    for idx, fi in enumerate(active_filters):
        if idx >= 4:
            break
        row_pos, lcol, vcol = positions[idx]
        h = headers[fi]

        ws.cell(row=row_pos, column=lcol, value=h).font = label_font
        ws.cell(row=row_pos, column=lcol).fill = light_fill
        ws.cell(row=row_pos, column=lcol).border = thin_border

        fcell = ws.cell(row=row_pos, column=vcol, value="Tümü")
        fcell.font = Font(name="Arial", bold=True, size=11, color=BLUE)
        fcell.fill = filter_fill
        fcell.border = thin_border
        fcell.alignment = Alignment(horizontal='center')

        # DataValidation from _Listeler
        lc, le = list_refs[fi]
        cl = get_column_letter(lc)
        dv = DataValidation(type="list", formula1=f"'_Listeler'!${cl}$1:${cl}${le}", allow_blank=True)
        dv.showInputMessage = True
        dv.showErrorMessage = True
        ws.add_data_validation(dv)
        dv.add(fcell)

        filter_cells.append((row_pos, vcol, fi))

    # Month filter
    month_filter_pos = None
    if has_month_filter:
        idx = len(active_filters)
        if idx < 4:
            row_pos, lcol, vcol = positions[idx]
            ws.cell(row=row_pos, column=lcol, value="Ay").font = label_font
            ws.cell(row=row_pos, column=lcol).fill = light_fill
            ws.cell(row=row_pos, column=lcol).border = thin_border

            fcell = ws.cell(row=row_pos, column=vcol, value="Tümü")
            fcell.font = Font(name="Arial", bold=True, size=11, color=BLUE)
            fcell.fill = filter_fill
            fcell.border = thin_border
            fcell.alignment = Alignment(horizontal='center')

            if month_list_ref:
                mlc, mle = month_list_ref
                mcl = get_column_letter(mlc)
                dv = DataValidation(type="list", formula1=f"'_Listeler'!${mcl}$1:${mcl}${mle}", allow_blank=True)
                dv.showInputMessage = True
                dv.showErrorMessage = True
                ws.add_data_validation(dv)
                dv.add(fcell)

            month_filter_pos = (row_pos, vcol)

    # VOLATILE HELPER CELLS (Row 7) — NOW() forces Excel recalc
    helper_refs = {}
    helper_col = 2
    for row_pos, vcol, fi in filter_cells:
        cell_ref = f"{get_column_letter(vcol)}{row_pos}"
        ws.cell(row=7, column=helper_col, value=f'=IF(NOW()>0,IF({cell_ref}="Tümü","*",{cell_ref}),"")')
        helper_refs[fi] = f"${get_column_letter(helper_col)}$7"
        helper_col += 2

    month_helper_ref = None
    if month_filter_pos:
        mcell_ref = f"{get_column_letter(month_filter_pos[1])}{month_filter_pos[0]}"
        ws.cell(row=7, column=helper_col, value=f'=IF(NOW()>0,IF({mcell_ref}="Tümü","*",{mcell_ref}),"")')
        month_helper_ref = f"${get_column_letter(helper_col)}$7"

    # Hide row 7
    ws.row_dimensions[7].height = 1
    for c in range(1, 9):
        ws.cell(row=7, column=c).font = Font(name="Arial", size=1, color="FFFFFF")

    # Build SUMIFS references
    val_col_letter = get_column_letter(main_value + 1)
    val_range = f"'Ham Veri'!{val_col_letter}$2:{val_col_letter}${de}"

    def build_criteria():
        """Build SUMIFS criteria pairs for Dashboard sheet"""
        parts = []
        for _, _, col_idx in filter_cells:
            data_col = get_column_letter(col_idx + 1)
            h_ref = helper_refs[col_idx]
            parts.append(f"'Ham Veri'!{data_col}$2:{data_col}${de},{h_ref}")
        if month_helper_ref and date_col is not None:
            m_col = get_column_letter(month_helper_col)
            parts.append(f"'Ham Veri'!{m_col}$2:{m_col}${de},{month_helper_ref}")
        return ",".join(parts)

    def build_x_criteria(extra_range=None, extra_val=None):
        """Build SUMIFS criteria for OTHER sheets (with Dashboard! prefix)"""
        parts = []
        for _, _, col_idx in filter_cells:
            data_col = get_column_letter(col_idx + 1)
            h_ref = helper_refs[col_idx]
            parts.append(f"'Ham Veri'!{data_col}$2:{data_col}${de},Dashboard!{h_ref}")
        if month_helper_ref and date_col is not None:
            m_col = get_column_letter(month_helper_col)
            parts.append(f"'Ham Veri'!{m_col}$2:{m_col}${de},Dashboard!{month_helper_ref}")
        if extra_range and extra_val:
            parts.append(f"{extra_range},{extra_val}")
        return ",".join(parts)

    criteria = build_criteria()

    # KPI Section (Row 8-10)
    r = 8
    for c in range(1, 9):
        ws.cell(row=r, column=c).fill = PatternFill("solid", fgColor=GREEN)
        ws.cell(row=r, column=c).font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    ws.cell(row=r, column=1, value="TEMEL METRİKLER (filtrelere göre otomatik güncellenir)")
    ws.row_dimensions[r].height = 28

    # Check for negative values (gelir-gider type)
    has_negatives = any(
        isinstance(r[main_value], (int, float)) and r[main_value] < 0
        for r in data_rows[:100] if main_value < len(r) and r[main_value]
    )

    kpi_formulas = [
        (1, "Toplam", f'=SUMIFS({val_range},{criteria})'),
        (3, "İşlem Sayısı", f'=COUNTIFS({criteria})'),
        (5, "Ortalama", '=IF(C10=0,0,A10/C10)'),
    ]

    if has_negatives:
        kpi_formulas.append((7, "Net Kar/Zarar", '=A10'))
    else:
        pivot_end = 13 + len(filter_data.get(filter_cols[0], [])) - 1 if filter_cols else 13
        kpi_formulas.append((7, "En Yüksek", f'=IF(C10=0,0,MAX(C13:C{pivot_end}))'))

    for _, (col, lbl, formula) in enumerate(kpi_formulas):
        ws.cell(row=9, column=col, value=lbl).font = label_font
        ws.cell(row=9, column=col).fill = light_fill
        ws.cell(row=9, column=col).border = thin_border
        cell = ws.cell(row=10, column=col, value=formula)
        cell.font = value_font
        cell.number_format = '#,##0'
        cell.border = thin_border

    # PIVOT TABLE: First categorical column
    formula_count = 0
    if filter_cols:
        pivot_col = filter_cols[0]
        pivot_vals = filter_data[pivot_col]
        pivot_header = headers[pivot_col]
        pivot_data_col = get_column_letter(pivot_col + 1)
        pivot_data_range = f"'Ham Veri'!{pivot_data_col}$2:{pivot_data_col}${de}"
        h_ref_first = helper_refs[pivot_col]

        r = 12
        for c in range(1, 6):
            ws.cell(row=r, column=c).fill = hdr_fill
            ws.cell(row=r, column=c).font = hdr_font
            ws.cell(row=r, column=c).border = thin_border
        ws.cell(row=r, column=1, value=pivot_header)
        ws.cell(row=r, column=2, value="İşlem Sayısı")
        ws.cell(row=r, column=3, value="Toplam Tutar")
        ws.cell(row=r, column=4, value="Ortalama")
        ws.cell(row=r, column=5, value="Pay %")
        ws.row_dimensions[r].height = 28

        # Build criteria WITHOUT the first filter
        other_parts = []
        for _, _, col_idx in filter_cells:
            if col_idx == pivot_col:
                continue
            data_col = get_column_letter(col_idx + 1)
            other_parts.append(f"'Ham Veri'!{data_col}$2:{data_col}${de},{helper_refs[col_idx]}")
        if month_helper_ref and date_col is not None:
            m_col = get_column_letter(month_helper_col)
            other_parts.append(f"'Ham Veri'!{m_col}$2:{m_col}${de},{month_helper_ref}")
        other_criteria = ("," + ",".join(other_parts)) if other_parts else ""

        for i, val in enumerate(pivot_vals):
            r = 13 + i
            safe = val.replace('"', '""')

            ws.cell(row=r, column=1, value=val).font = bold_sm
            ws.cell(row=r, column=1).border = thin_border

            # IF(OR(filter="*", filter=value), COUNTIFS/SUMIFS, 0)
            ws.cell(row=r, column=2,
                    value=f'=IF(OR({h_ref_first}="*",{h_ref_first}="{safe}"),COUNTIFS({pivot_data_range},"{safe}"{other_criteria}),0)')
            ws.cell(row=r, column=2).font = normal_font
            ws.cell(row=r, column=2).number_format = '#,##0'
            ws.cell(row=r, column=2).border = thin_border
            formula_count += 1

            ws.cell(row=r, column=3,
                    value=f'=IF(OR({h_ref_first}="*",{h_ref_first}="{safe}"),SUMIFS({val_range},{pivot_data_range},"{safe}"{other_criteria}),0)')
            ws.cell(row=r, column=3).font = bold_sm
            ws.cell(row=r, column=3).number_format = '#,##0'
            ws.cell(row=r, column=3).border = thin_border
            formula_count += 1

            ws.cell(row=r, column=4, value=f'=IF(B{r}=0,0,C{r}/B{r})')
            ws.cell(row=r, column=4).font = normal_font
            ws.cell(row=r, column=4).number_format = '#,##0'
            ws.cell(row=r, column=4).border = thin_border
            formula_count += 1

            if i % 2 == 0:
                for c in range(1, 6):
                    ws.cell(row=r, column=c).fill = alt_fill

        # Total row
        total_r = 13 + len(pivot_vals)
        first_r = 13
        last_r = total_r - 1
        for c in range(1, 6):
            ws.cell(row=total_r, column=c).fill = blue_fill
            ws.cell(row=total_r, column=c).font = Font(name="Arial", bold=True, color="FFFFFF", size=10)
            ws.cell(row=total_r, column=c).border = thin_border
        ws.cell(row=total_r, column=1, value="TOPLAM")
        ws.cell(row=total_r, column=2, value=f'=SUM(B{first_r}:B{last_r})').number_format = '#,##0'
        ws.cell(row=total_r, column=3, value=f'=SUM(C{first_r}:C{last_r})').number_format = '#,##0'
        ws.cell(row=total_r, column=4, value=f'=IF(B{total_r}=0,0,C{total_r}/B{total_r})').number_format = '#,##0'
        ws.cell(row=total_r, column=5, value=1).number_format = '0%'
        formula_count += 4

        for i in range(len(pivot_vals)):
            r = 13 + i
            ws.cell(row=r, column=5, value=f'=IF(C${total_r}=0,0,C{r}/ABS(C${total_r}))')
            ws.cell(row=r, column=5).number_format = '0.0%'
            ws.cell(row=r, column=5).border = thin_border
            formula_count += 1

        # Data bars
        db = DataBarRule(start_type='min', end_type='max', color=BLUE)
        ws.conditional_formatting.add(f'E{first_r}:E{last_r}', db)

        # Bar chart
        chart1 = BarChart()
        chart1.type = "col"
        chart1.title = f"{pivot_header} Bazında Analiz"
        chart1.y_axis.title = headers[main_value]
        chart1.style = 10
        chart1.width = 18
        chart1.height = 12
        data_ref = Reference(ws, min_col=3, min_row=12, max_row=last_r)
        cats_ref = Reference(ws, min_col=1, min_row=first_r, max_row=last_r)
        chart1.add_data(data_ref, titles_from_data=True)
        chart1.set_categories(cats_ref)
        chart1.shape = 4
        s = chart1.series[0]
        s.graphicalProperties.solidFill = BLUE
        ws.add_chart(chart1, f"A{total_r + 2}")

        # Pie chart
        chart2 = PieChart()
        chart2.title = f"{pivot_header} Dağılımı"
        chart2.style = 10
        chart2.width = 14
        chart2.height = 12
        chart2.add_data(data_ref, titles_from_data=True)
        chart2.set_categories(cats_ref)
        chart2.dataLabels = DataLabelList()
        chart2.dataLabels.showPercent = True
        ws.add_chart(chart2, f"E{total_r + 2}")

    # Column widths
    for c, w in enumerate([22, 16, 18, 16, 14, 14, 14, 14], 1):
        ws.column_dimensions[get_column_letter(c)].width = w

    # --- SECOND PIVOT SHEET ---
    if len(filter_cols) >= 2:
        pivot2_col = filter_cols[1]
        pivot2_vals = filter_data[pivot2_col]
        pivot2_header = headers[pivot2_col]
        pivot2_data_col = get_column_letter(pivot2_col + 1)
        pivot2_range = f"'Ham Veri'!{pivot2_data_col}$2:{pivot2_data_col}${de}"

        ws2 = wb.create_sheet(f"{pivot2_header} Pivot")
        ws2.sheet_properties.tabColor = PURPLE
        ws2.merge_cells('A1:F1')
        ws2['A1'] = f"ExcelMind — {pivot2_header} Bazlı Analiz"
        ws2['A1'].font = Font(name="Arial", bold=True, color=PURPLE, size=16)
        ws2.row_dimensions[1].height = 36

        r = 4
        for c in range(1, 6):
            ws2.cell(row=r, column=c).fill = PatternFill("solid", fgColor=PURPLE)
            ws2.cell(row=r, column=c).font = hdr_font
            ws2.cell(row=r, column=c).border = thin_border
        ws2.cell(row=r, column=1, value=pivot2_header)
        ws2.cell(row=r, column=2, value="İşlem Sayısı")
        ws2.cell(row=r, column=3, value="Toplam Tutar")
        ws2.cell(row=r, column=4, value="Ortalama")
        ws2.cell(row=r, column=5, value="Pay %")

        x_crit = build_x_criteria()

        for i, val in enumerate(pivot2_vals):
            r = 5 + i
            safe = val.replace('"', '""')
            ws2.cell(row=r, column=1, value=val).font = bold_sm
            ws2.cell(row=r, column=1).border = thin_border
            ws2.cell(row=r, column=2, value=f'=COUNTIFS({pivot2_range},"{safe}",{x_crit})').number_format = '#,##0'
            ws2.cell(row=r, column=2).font = normal_font
            ws2.cell(row=r, column=2).border = thin_border
            ws2.cell(row=r, column=3, value=f'=SUMIFS({val_range},{pivot2_range},"{safe}",{x_crit})').number_format = '#,##0'
            ws2.cell(row=r, column=3).font = bold_sm
            ws2.cell(row=r, column=3).border = thin_border
            ws2.cell(row=r, column=4, value=f'=IF(B{r}=0,0,C{r}/B{r})').number_format = '#,##0'
            ws2.cell(row=r, column=4).font = normal_font
            ws2.cell(row=r, column=4).border = thin_border
            formula_count += 4

            if i % 2 == 0:
                for c in range(1, 6):
                    ws2.cell(row=r, column=c).fill = alt_fill

        p2_total = 5 + len(pivot2_vals)
        for c in range(1, 6):
            ws2.cell(row=p2_total, column=c).fill = PatternFill("solid", fgColor=PURPLE)
            ws2.cell(row=p2_total, column=c).font = Font(name="Arial", bold=True, color="FFFFFF", size=10)
            ws2.cell(row=p2_total, column=c).border = thin_border
        ws2.cell(row=p2_total, column=1, value="TOPLAM")
        ws2.cell(row=p2_total, column=2, value=f'=SUM(B5:B{p2_total - 1})').number_format = '#,##0'
        ws2.cell(row=p2_total, column=3, value=f'=SUM(C5:C{p2_total - 1})').number_format = '#,##0'
        ws2.cell(row=p2_total, column=4, value=f'=IF(B{p2_total}=0,0,C{p2_total}/B{p2_total})').number_format = '#,##0'
        formula_count += 3

        for i in range(len(pivot2_vals)):
            r = 5 + i
            ws2.cell(row=r, column=5, value=f'=IF(C${p2_total}=0,0,C{r}/ABS(C${p2_total}))').number_format = '0.0%'
            ws2.cell(row=r, column=5).border = thin_border
            formula_count += 1

        # Chart
        pchart = BarChart()
        pchart.type = "bar"
        pchart.title = f"{pivot2_header} Bazında Analiz"
        pchart.style = 10
        pchart.width = 22
        pchart.height = 14
        pdata = Reference(ws2, min_col=3, min_row=4, max_row=p2_total - 1)
        pcats = Reference(ws2, min_col=1, min_row=5, max_row=p2_total - 1)
        pchart.add_data(pdata, titles_from_data=True)
        pchart.set_categories(pcats)
        s = pchart.series[0]
        s.graphicalProperties.solidFill = PURPLE
        ws2.add_chart(pchart, f"A{p2_total + 2}")

        for c, w in enumerate([28, 16, 20, 16, 12], 1):
            ws2.column_dimensions[get_column_letter(c)].width = w

    # --- MONTHLY TREND SHEET ---
    if months_list and date_col is not None:
        ws_trend = wb.create_sheet("Aylık Trend")
        ws_trend.sheet_properties.tabColor = GREEN
        ws_trend.merge_cells('A1:E1')
        ws_trend['A1'] = "ExcelMind — Aylık Trend"
        ws_trend['A1'].font = Font(name="Arial", bold=True, color=GREEN, size=16)
        ws_trend.row_dimensions[1].height = 36

        r = 4
        for c in range(1, 5):
            ws_trend.cell(row=r, column=c).fill = PatternFill("solid", fgColor=GREEN)
            ws_trend.cell(row=r, column=c).font = hdr_font
            ws_trend.cell(row=r, column=c).border = thin_border
        ws_trend.cell(row=r, column=1, value="Dönem")
        ws_trend.cell(row=r, column=2, value="Toplam")
        ws_trend.cell(row=r, column=3, value="Değişim")
        ws_trend.cell(row=r, column=4, value="Değişim %")

        m_col_letter = get_column_letter(month_helper_col)
        m_range = f"'Ham Veri'!{m_col_letter}$2:{m_col_letter}${de}"
        x_crit_month = build_x_criteria()

        for i, ml in enumerate(months_list):
            r = 5 + i
            safe_m = ml.replace('"', '""')

            ws_trend.cell(row=r, column=1, value=ml).font = bold_sm
            ws_trend.cell(row=r, column=1).border = thin_border
            ws_trend.cell(row=r, column=2,
                          value=f'=SUMIFS({val_range},{m_range},"{safe_m}",{x_crit_month})').number_format = '#,##0'
            ws_trend.cell(row=r, column=2).font = value_font
            ws_trend.cell(row=r, column=2).border = thin_border
            formula_count += 1

            if i > 0:
                ws_trend.cell(row=r, column=3, value=f'=B{r}-B{r - 1}').number_format = '#,##0'
                ws_trend.cell(row=r, column=3).border = thin_border
                ws_trend.cell(row=r, column=4, value=f'=IF(B{r - 1}=0,0,(B{r}-B{r - 1})/ABS(B{r - 1}))').number_format = '0.0%'
                ws_trend.cell(row=r, column=4).border = thin_border
                formula_count += 2
            else:
                ws_trend.cell(row=r, column=3, value="-").border = thin_border
                ws_trend.cell(row=r, column=4, value="-").border = thin_border

            if i % 2 == 0:
                for c in range(1, 5):
                    ws_trend.cell(row=r, column=c).fill = alt_fill

        # Conditional formatting
        t_last = 4 + len(months_list)
        green_cf = CellIsRule(operator='greaterThan', formula=['0'], font=Font(color=GREEN), fill=green_fill)
        red_cf = CellIsRule(operator='lessThan', formula=['0'], font=Font(color=RED),
                            fill=PatternFill("solid", fgColor="FDEDEC"))
        ws_trend.conditional_formatting.add(f'C6:D{t_last}', green_cf)
        ws_trend.conditional_formatting.add(f'C6:D{t_last}', red_cf)

        # Line chart
        lchart = LineChart()
        lchart.title = "Aylık Trend"
        lchart.y_axis.title = headers[main_value]
        lchart.style = 10
        lchart.width = 24
        lchart.height = 14
        ldata = Reference(ws_trend, min_col=2, min_row=4, max_row=t_last)
        lcats = Reference(ws_trend, min_col=1, min_row=5, max_row=t_last)
        lchart.add_data(ldata, titles_from_data=True)
        lchart.set_categories(lcats)
        s = lchart.series[0]
        s.graphicalProperties.line.solidFill = GREEN
        s.graphicalProperties.line.width = 28000
        s.smooth = True
        ws_trend.add_chart(lchart, f"A{t_last + 2}")

        for c, w in enumerate([20, 20, 16, 14], 1):
            ws_trend.column_dimensions[get_column_letter(c)].width = w

    # --- Finalize ---
    wb.move_sheet("Dashboard", offset=-10)
    wb.calculation = openpyxl.workbook.properties.CalcProperties(fullCalcOnLoad=True)

    wb.save(output_path)

    stats = {
        'sheets': wb.sheetnames,
        'formulas': formula_count,
        'filters': len(filter_cols),
        'rows': n_rows,
        'columns': len(headers),
        'main_value': str(headers[main_value]),
        'has_month_filter': bool(months_list),
    }

    print(f"[ENGINE] Smart Excel created: {output_path}")
    print(f"[ENGINE] Sheets: {wb.sheetnames}, Formulas: {formula_count}")

    return stats
